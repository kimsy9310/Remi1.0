"""
Warm-loop learning: fit a data-driven response model from real prototype data,
and COMPARE two predictors (PLS vs GP) by leave-one-out cross-validation.

Why these two:
  - PLS (NIPALS, multi-response): built for small-n / many correlated ingredients
    (p>n) and correlated responses -> absorbs attribute-attribute correlation,
    interpretable. Linear by default.
  - GP (RBF): nonlinear, gives predictive uncertainty (feeds BO), but data-hungry
    and prior-sensitive at small n.

Pure numpy (sandbox has no sklearn). Reads the warm-loop Excel template.
"""
from __future__ import annotations
import numpy as np


# ----------------------------- data loading -----------------------------
def load_dataset(xlsx_path):
    """Return dict with X (recipes), Y (mean sensory, benchmark-relative), names,
    replicate noise, and optional instrumental block. Aligns by sample_id."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)

    def sheet_rows(name):
        ws = wb[name]; rows = list(ws.iter_rows(values_only=True))
        hdr = [h for h in rows[0]]
        data = [r for r in rows[1:] if r and r[0] not in (None, "")]
        return hdr, data

    rh, rd = sheet_rows("recipes")
    ing_idx = [i for i, h in enumerate(rh) if str(h).startswith("ing__")]
    sid_i = rh.index("sample_id")
    recipes = {}
    for r in rd:
        recipes[r[sid_i]] = np.array([_num(r[i]) for i in ing_idx], float)
    ing_names = [rh[i] for i in ing_idx]

    sh, sd = sheet_rows("sensory")
    sens_idx = [i for i, h in enumerate(sh) if str(h).startswith("sens__")]
    ssid = sh.index("sample_id")
    sens = {}
    for r in sd:
        sid = r[ssid]
        vec = [_num(r[i]) for i in sens_idx]
        sens.setdefault(sid, []).append(vec)
    sens_names = [sh[i] for i in sens_idx]

    ids = [s for s in recipes if s in sens]
    X = np.array([recipes[s] for s in ids])
    Y = np.array([np.nanmean(np.array(sens[s], float), axis=0) for s in ids])
    # replicate noise (std across reps, averaged) -> per-attribute measurement noise
    rep_noise = np.zeros(Y.shape[1])
    cnt = 0
    for s in ids:
        a = np.array(sens[s], float)
        if a.shape[0] > 1:
            rep_noise += np.nanstd(a, axis=0, ddof=1); cnt += 1
    rep_noise = rep_noise / cnt if cnt else None
    return {"ids": ids, "X": X, "Y": Y, "ing_names": ing_names,
            "sens_names": sens_names, "rep_noise": rep_noise}

def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return np.nan


# ----------------------------- PLS (NIPALS, PLS2) -----------------------------
class PLS:
    def __init__(self, n_components=2, scale=True):
        self.a = n_components; self.scale = scale
    def fit(self, X, Y):
        X = np.asarray(X, float); Y = np.asarray(Y, float)
        self.xm = np.nanmean(X, 0); self.ym = np.nanmean(Y, 0)
        Xs = np.nan_to_num(X - self.xm); Ys = np.nan_to_num(Y - self.ym)
        self.xsd = Xs.std(0) + 1e-9; self.ysd = Ys.std(0) + 1e-9
        if self.scale:
            Xs = Xs / self.xsd; Ys = Ys / self.ysd
        n, p = Xs.shape; m = Ys.shape[1]
        W = np.zeros((p, self.a)); P = np.zeros((p, self.a)); C = np.zeros((m, self.a)); T = np.zeros((n, self.a))
        E, F = Xs.copy(), Ys.copy()
        for k in range(self.a):
            u = F[:, 0].copy()
            for _ in range(200):
                w = E.T @ u; w /= (np.linalg.norm(w) + 1e-12)
                t = E @ w
                c = F.T @ t / (t @ t + 1e-12)
                u_new = F @ c / (c @ c + 1e-12)
                if np.linalg.norm(u_new - u) < 1e-9: u = u_new; break
                u = u_new
            pl = E.T @ t / (t @ t + 1e-12)
            E -= np.outer(t, pl); F -= np.outer(t, c)
            W[:, k] = w; P[:, k] = pl; C[:, k] = c; T[:, k] = t
        self.B = W @ np.linalg.pinv(P.T @ W) @ C.T   # in scaled space
        return self
    def predict(self, X):
        Xs = np.nan_to_num(np.asarray(X, float) - self.xm)
        if self.scale: Xs = Xs / self.xsd
        Ys = Xs @ self.B
        if self.scale: Ys = Ys * self.ysd
        return Ys + self.ym


# ----------------------------- GP (RBF, per response) -----------------------------
class GPR:
    def __init__(self, l=0.8, sf=1.0, noise=0.25):
        self.l, self.sf, self.noise = l, sf, noise
    def _k(self, A, B):
        d = ((A[:, None, :] - B[None, :, :]) ** 2).sum(-1)
        return self.sf**2 * np.exp(-0.5 * d / self.l**2)
    def fit(self, X, y):
        self.xm = X.mean(0); self.xs = X.std(0) + 1e-9
        self.Xn = (X - self.xm) / self.xs
        self.ym = y.mean(); self.ys = y.std() + 1e-9
        yc = (y - self.ym) / self.ys
        K = self._k(self.Xn, self.Xn) + self.noise**2 * np.eye(len(X))
        self.L = np.linalg.cholesky(K)
        self.al = np.linalg.solve(self.L.T, np.linalg.solve(self.L, yc)); return self
    def predict(self, X):
        Xn = (X - self.xm) / self.xs; Ks = self._k(Xn, self.Xn)
        return (Ks @ self.al) * self.ys + self.ym


# ----------------------------- LOO comparison -----------------------------
def loo_compare(X, Y, pls_components=(1, 2, 3), gp_len=0.8):
    """Leave-one-out CV. Returns per-response RMSE & Pearson r for PLS(best a) and GP."""
    n, m = Y.shape
    out = {}
    # PLS: pick best #components by overall LOO RMSE
    best = None
    for a in pls_components:
        if a >= n: continue
        pred = np.zeros_like(Y)
        for i in range(n):
            tr = [j for j in range(n) if j != i]
            pred[i] = PLS(a).fit(X[tr], Y[tr]).predict(X[i:i+1])[0]
        rmse = np.sqrt(np.nanmean((pred - Y) ** 2))
        if best is None or rmse < best[0]:
            best = (rmse, a, pred)
    pls_pred = best[2]; out["pls_components"] = best[1]
    # GP per response
    gp_pred = np.zeros_like(Y)
    for i in range(n):
        tr = [j for j in range(n) if j != i]
        for c in range(m):
            gp_pred[i, c] = GPR(l=gp_len).fit(X[tr], Y[tr, c]).predict(X[i:i+1])[0]
    def metrics(P):
        rmse = np.sqrt(np.nanmean((P - Y) ** 2, axis=0))
        r = np.array([_corr(P[:, c], Y[:, c]) for c in range(m)])
        return rmse, r
    out["pls_rmse"], out["pls_r"] = metrics(pls_pred)
    out["gp_rmse"], out["gp_r"] = metrics(gp_pred)
    out["pls_pred"], out["gp_pred"] = pls_pred, gp_pred
    return out

def _corr(a, b):
    a = np.asarray(a); b = np.asarray(b)
    if a.std() < 1e-9 or b.std() < 1e-9: return np.nan
    return float(np.corrcoef(a, b)[0, 1])
