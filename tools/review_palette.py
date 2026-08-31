# -*- coding: utf-8 -*-
"""
팔레트 검토표 생성 — 범위와 등급을 한 장에서 판정할 수 있게.

왜 필요한가
-----------
팔레트 범위 62종이 다 채워졌지만 **전부 초안**이다. 아이스크림 50종은 표준
사용량에서, 쌀 우유 12종은 실측 사용폭에서 넣었다. 온톨로지의 limitations 에서
뽑으려 했으나 39종 중 숫자가 있는 것은 잔탄검 하나뿐이었다.

앱 ④ 팔레트 탭에서도 고칠 수 있지만, 62종을 훑기에는 표가 낫다. DRAFT 검토표와
같은 방식으로 판정 열을 둔다.

정렬
----
검토 우선순위가 높은 것부터 온다.
  1. 범위가 넓은 것 — 상/하한 비가 클수록 모형이 크게 흔들 수 있다
  2. 담당축이 많은 것 — 여러 축에 영향을 주므로 틀리면 파급이 크다
  3. 근거가 약한 것 — '표준 사용량(초안)' 은 문헌값이고 실측이 아니다

실행:
    python tools/review_palette.py
    -> data/palette_review.xlsx
"""
from __future__ import annotations

import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
sys.path.insert(0, os.path.join(_ROOT, "app"))
sys.path.insert(0, os.path.join(_ROOT, "engine"))

import palette                                   # noqa: E402
from formulator.v2adapter import V2Ontology      # noqa: E402

OUT = os.path.join(_ROOT, "data", "palette_review.xlsx")

COLS = ["프로파일", "슬롯", "재료", "온톨로지ID", "등급",
        "하한", "상한", "폭(배)", "담당축수", "담당축",
        "범위근거", "판정", "고칠 하한", "고칠 상한", "메모"]

GRADE_FILL = {"필수": "CFE2F3", "권장": "D9EAD3", "옵션": "FFF2CC",
              "제한": "FCE5CD", "제외": "EFEFEF"}


def span(lo, hi):
    """상/하한 비. 하한이 0 이면 폭을 비로 잴 수 없으므로 None."""
    if lo is None or hi is None:
        return None
    if lo <= 1e-9:
        return None
    return hi / lo


def main():
    onto = V2Ontology()
    rows = []
    for prof in palette.profiles_in():
        t = palette.load(prof)
        try:
            cards = onto._ref.load_cards(prof, onto.layers)
            core = [c["term_id"] for c in cards
                    if c["tier"] == "core" and c["evidence_required"] != "sample_aged"]
        except Exception:                                      # noqa: BLE001
            core = []
        for r in t.rows:
            if r["등급"] == "제외":
                continue                       # 안 쓰는 재료는 범위가 필요 없다
            ax = (r["담당축"] or "").strip()
            n_ax = len(ax.split()) if ax else 0
            sp = span(r["하한"], r["상한"])
            rows.append(dict(
                프로파일=prof, 슬롯=r["슬롯"], 재료=r["재료"],
                온톨로지ID=r["온톨로지ID"], 등급=r["등급"],
                하한=r["하한"], 상한=r["상한"],
                폭=sp, 담당축수=n_ax, 담당축=ax,
                범위근거=r["범위근거"]))

    # 검토 우선순위: 폭 넓은 것 · 담당축 많은 것 먼저
    rows.sort(key=lambda d: (-(d["폭"] or 0), -d["담당축수"], d["프로파일"]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "palette_review"
    ws.append(COLS)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E7EBE5")
    for d in rows:
        ws.append([d["프로파일"], d["슬롯"], d["재료"], d["온톨로지ID"], d["등급"],
                   d["하한"], d["상한"],
                   round(d["폭"], 1) if d["폭"] else "",
                   d["담당축수"] or "", d["담당축"], d["범위근거"], "", "", "", ""])
    gi = COLS.index("등급") + 1
    for i in range(2, ws.max_row + 1):
        g = ws.cell(i, gi).value
        if g in GRADE_FILL:
            f = PatternFill("solid", fgColor=GRADE_FILL[g])
            for c in range(1, len(COLS) + 1):
                ws.cell(i, c).fill = f
    for i, w in enumerate([20, 14, 22, 32, 7, 9, 9, 8, 8, 30, 46, 9, 10, 10, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws2 = wb.create_sheet("README")
    for line in [
        "팔레트 검토표",
        "",
        "생성: tools/review_palette.py",
        "",
        "■ 이 값들은 전부 초안입니다",
        "  아이스크림 50종 — 아이스크림 표준 사용량에서",
        "  쌀 우유 12종  — 실측 18건의 사용 폭에서",
        "  온톨로지의 limitations 에서 뽑으려 했으나 39종 중 숫자가 있는 것은",
        "  잔탄검 하나뿐이었습니다(\"> ~0.5% -> slimy/stringy\").",
        "",
        "■ 이 값이 무엇을 결정하나",
        "  1) 제안의 상·하한 — 이 밖으로는 절대 안 나갑니다",
        "  2) 사전계수의 스케일 — 폭이 곧 '1 표준편차' 의 크기가 됩니다",
        "     범위를 두 배 넓게 잡으면 모형이 그 재료의 효과를 절반으로 봅니다",
        "",
        "■ 정렬 순서",
        "  폭(상한/하한 비)이 큰 것부터. 폭이 넓을수록 모형이 크게 흔들 수 있어",
        "  틀렸을 때 파급이 큽니다. 그다음이 담당축이 많은 재료입니다.",
        "  하한이 0 인 재료는 비를 낼 수 없어 폭이 비어 있습니다(선택 재료).",
        "",
        "■ 하한이 0 이라는 것의 뜻",
        "  '안 넣어도 된다' 입니다. 모든 배합에 반드시 들어가야 하는 재료만",
        "  하한을 0 보다 크게 둡니다(DoE 규칙 3). 실제로 이걸 잘못 잡아서,",
        "  18건 중 4건에만 쓰인 해바라기유에 하한 5.8 이 걸려 모든 제안에",
        "  강제로 들어간 적이 있습니다.",
        "",
        "■ 판정 열에 적을 것",
        "  OK      그대로 간다",
        "  수정    '고칠 하한' / '고칠 상한' 에 값을 적는다",
        "  제외    이 재료를 팔레트에서 뺀다 (등급을 '제외' 로 바꿉니다)",
        "  보류    실측을 보고 정한다",
        "",
        "■ 판정 후",
        "  이 파일을 그대로 두시면 제가 읽어 palette.xlsx 에 반영합니다.",
        "  반영 후 제안이 범위 안에 머무는지 검사를 돌립니다.",
    ]:
        ws2.append([line])
    ws2.column_dimensions["A"].width = 96

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)

    import collections
    byp = collections.Counter(d["프로파일"] for d in rows)
    wide = [d for d in rows if (d["폭"] or 0) >= 5]
    zero = [d for d in rows if d["하한"] is not None and d["하한"] <= 1e-9]
    print(f"작성: {OUT}")
    print(f"  검토 항목 {len(rows)}건 {dict(byp)}")
    print(f"  폭 5배 이상 {len(wide)}건 (먼저 보실 것)")
    print(f"  하한 0 인 선택 재료 {len(zero)}건")


if __name__ == "__main__":
    main()
