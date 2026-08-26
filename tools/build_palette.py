# -*- coding: utf-8 -*-
"""
팔레트 표 생성기 — 사람이 쓴 shortlist 를 기계가 읽는 표로 바꾼다.

왜 변환이 필요한가
------------------
기존 `icecream_palette_shortlist.xlsx` 는 잘 만들어진 큐레이션이지만 사람 눈을
전제로 쓰였다. 기계가 읽으려면 세 가지가 걸린다.

  1. 한 줄에 ID 가 여러 개다.  "ING.coconut_oil / ING.refined_coconut_oil"
     게다가 표기가 섞여 있다.  "ING.sucrose / glucose_syrup / dextrose"
                               "ING.citric_acid / ..._anhydrous"
  2. 기능군이 헤더 행으로만 있고 각 재료 행에는 없다.
  3. 용량 범위가 없다. 이게 없으면 제안이 물리적으로 불가능한 값을 낸다(G5).

여기서 ①②를 풀고 ③의 자리를 만든다. 값 자체는 사람이 채운다 — 도메인 지식이다.

**추측하지 않는다.** 풀어낸 ID 는 전부 온톨로지에 대조하고, 못 찾은 것은
'확인' 열에 남겨 사람이 고치게 한다. 조용히 비슷한 걸 갖다 붙이면 그게 더 나쁘다.

실행:
    python tools/build_palette.py
"""
from __future__ import annotations

import os
import re
import sys

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
sys.path.insert(0, os.path.join(_ROOT, "engine"))
sys.path.insert(0, os.path.join(_ROOT, "app"))

from formulator.v2adapter import V2Ontology            # noqa: E402
import dataio                                          # noqa: E402

SRC = os.path.join(
    os.path.dirname(os.path.dirname(_ROOT)),
    "Claude", "recovered_food_recipe_app", "_정리_remi_20260724",
    "03_아이스크림", "현재", "icecream_palette_shortlist.xlsx")
OUT = os.path.join(_ROOT, "data", "palette.xlsx")

COLS = ["프로파일", "슬롯", "재료", "온톨로지ID", "등급",
        "하한", "상한", "범위근거", "담당축", "메모", "확인"]

GRADE_FILL = {
    "필수": "CFE2F3",   # 파랑
    "권장": "D9EAD3",   # 초록
    "옵션": "FFF2CC",   # 노랑
    "제한": "FCE5CD",   # 주황
    "제외": "EFEFEF",   # 회색
}

# HANDOFF §3.3 이 적어 둔 iter0 DoE 레버 범위. 근거가 있는 것만 씨앗으로 넣는다.
SEED_ICECREAM = {
    "ING.rice_extract":        (12.0, 14.0, "HANDOFF §3.3 · 필수 base 13% 고정 부근"),
    "ING.refined_coconut_oil": (6.0, 14.0, "HANDOFF §3.3 · iter0 레버 정제야자유"),
    "ING.allulose":            (9.0, 15.0, "HANDOFF §3.3 · iter0 레버 알룰로스"),
    "ING.guar_gum":            (0.15, 0.35, "HANDOFF §3.3 · iter0 레버 구아검"),
}


def parse_ids(raw):
    """
    'ING.a / b / ...c_suffix' → ['ING.a', 'ING.b', ...]

    표기가 세 가지 섞여 있다.
      완전형    ING.lambda_carrageenan
      접두생략  glucose_syrup            → ING. 를 붙인다
      말줄임    ..._anhydrous            → 앞 항목의 어간에 이어 붙인다
                ...artificial            (구분자 없는 형태도 있다)
    말줄임은 원문이 무엇을 뜻했는지 확실하지 않으므로 후보를 만들되
    온톨로지 대조에서 걸러지게 둔다.
    """
    if not raw:
        return []
    out, prev = [], None
    for tok in str(raw).split("/"):
        t = tok.strip()
        if not t:
            continue
        if t.startswith("..."):
            tail = t[3:]
            if prev:
                base = prev[len("ING."):]
                stem = base.rsplit("_", 1)[0] if not tail.startswith("_") else base
                cand = f"ING.{stem}{tail}" if tail.startswith("_") else f"ING.{stem}_{tail}"
                out.append(cand.replace("__", "_"))
            continue
        if not t.startswith("ING."):
            t = "ING." + t
        out.append(t)
        prev = t
    return out


def axes_of(onto, profile, ing_id, core_terms):
    """이 재료가 이 프로파일에서 움직이는 목표축. 규칙 3(슬롯마다 담당축)의 자동 부분."""
    eff = onto.effects_for(profile).get(ing_id) or {}
    bits = []
    for t in core_terms:
        if t in eff:
            arrow = "↑" if eff[t]["sign"] > 0 else "↓"
            bits.append(f"{t.split('.')[-1]}{arrow}")
    return " ".join(bits)


def read_shortlist(path):
    """기능군 헤더를 각 행에 내려 붙이며 읽는다."""
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    rows, slot = [], None
    for r in ws.iter_rows(values_only=True):
        c = ["" if x is None else str(x).strip() for x in list(r) + [""] * 5][:5]
        group, name, ids, grade, memo = c
        if group and not name and not ids:
            slot = group                     # 기능군 헤더 행
            continue
        if name and grade:
            # 표 헤더 행("기능군 | 재료 | 온톨로지 ID | 등급 | ...")이 여기 걸린다.
            # name='재료', grade='등급' 이라 값 행과 모양이 같으므로 값으로 판별한다.
            if grade not in GRADE_FILL:
                continue
            rows.append(dict(slot=slot or "(미분류)", name=name, ids=ids,
                             grade=grade, memo=memo))
    return rows


def build():
    onto = V2Ontology()
    records, unresolved = [], []

    # ---------------------------------------------------- 아이스크림 (shortlist)
    prof = "icecream"
    cards = onto._ref.load_cards(prof, onto.layers)
    core = [c["term_id"] for c in cards
            if c["tier"] == "core" and c["evidence_required"] != "sample_aged"]

    for row in read_shortlist(SRC):
        ids = parse_ids(row["ids"])
        if not ids:
            unresolved.append((row["name"], row["ids"], "ID 를 하나도 못 읽음"))
            continue
        for gid in ids:
            ok = gid in onto.ingredients
            if not ok:
                unresolved.append((row["name"], gid, "온톨로지에 없음"))
            lo, hi, basis = SEED_ICECREAM.get(gid, ("", "", ""))
            records.append(dict(
                프로파일=prof, 슬롯=row["slot"], 재료=row["name"], 온톨로지ID=gid,
                등급=row["grade"], 하한=lo, 상한=hi, 범위근거=basis,
                담당축=axes_of(onto, prof, gid, core) if ok else "",
                메모=row["memo"],
                확인="" if ok else "ID 확인 필요"))

    # ---------------------------------------------------- 쌀 우유 (실측에서 씨앗)
    prof2 = "beverage_rice_milk"
    try:
        data = dataio.load_warmloop(
            os.path.join(_ROOT, "data", "beverage_warmloop_260720.xlsx"), onto, prof2)
        cards2 = onto._ref.load_cards(prof2, onto.layers)
        core2 = [c["term_id"] for c in cards2
                 if c["tier"] == "core" and c["evidence_required"] != "sample_aged"]
        filler = onto.filler_of(prof2)
        for i, gid in enumerate(data.names):
            if gid == filler:
                continue
            col = data.X[:, i]
            used = col[col > 0]
            if used.size == 0:
                continue
            # **하한은 그 재료가 모든 샘플에 들어갔을 때만 0 보다 크다.**
            # 일부 샘플에만 쓰인 재료에 사용량 최솟값을 하한으로 걸면, 그 재료를
            # 안 쓰던 배합에까지 강제로 밀어 넣는다. 실제로 그렇게 됐다 —
            # 해바라기유가 18건 중 4건에만 쓰였는데 하한 5.8 이 걸려, 기준 배합
            # (해바라기유 0)에서 제안을 내면 곧장 +5.8 이 튀어나와 다른 모든
            # 변화를 덮어버렸다. DoE 규칙 3 이 말하는 '필수 기능 재료' 만
            # 0 을 피해야 하고, 선택 재료는 0 이 정상이다.
            always = int((col > 0).sum()) == len(col)
            lo = float(used.min()) if always else 0.0
            hi = float(used.max())
            if always and hi - lo < 1e-9:
                # 모든 런에서 같은 수준으로만 썼다 → 범위에 대해 아는 것이 없다.
                # 작업점 둘레로 임의 폭을 주되 그렇게 적어 둔다. 상한=하한은
                # propose 의 경계로 못 쓰고(구간이 비어 있다) 스케일도 0 이 된다.
                lo, hi = lo * 0.5, hi * 1.5
                basis = f"실측 {used.size}건 모두 같은 수준({used[0]:.3g}) — 임의 폭, 검토 필요"
            elif always:
                basis = f"실측 {used.size}건 전부에 사용 — 사용 범위, 검토 필요"
            else:
                basis = (f"실측 {len(col)}건 중 {used.size}건에만 사용 — 선택 재료라 "
                         f"하한 0, 상한은 최대 사용량. 검토 필요")
            records.append(dict(
                프로파일=prof2, 슬롯="(실측에서 자동)", 재료=gid.replace("ING.", ""),
                온톨로지ID=gid, 등급="권장",
                하한=round(lo, 4), 상한=round(hi, 4),
                범위근거=basis,
                담당축=axes_of(onto, prof2, gid, core2),
                메모="", 확인="슬롯·등급 확인 필요"))
    except Exception as e:                                    # noqa: BLE001
        print(f"  (쌀 우유 씨앗 생략: {e})")

    # ---------------------------------------------------- 쓰기
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "palette"
    ws.append(COLS)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E7EBE5")
    for rec in records:
        ws.append([rec[c] for c in COLS])
    gi = COLS.index("등급") + 1
    for r in range(2, ws.max_row + 1):
        g = ws.cell(r, gi).value
        if g in GRADE_FILL:
            f = PatternFill("solid", fgColor=GRADE_FILL[g])
            for c in range(1, len(COLS) + 1):
                ws.cell(r, c).fill = f
    widths = [20, 14, 24, 34, 7, 8, 8, 34, 30, 46, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws2 = wb.create_sheet("README")
    for line in [
        "팔레트 표 — 이 파일이 앱의 재료 범위 원본이다",
        "",
        "생성: tools/build_palette.py (icecream_palette_shortlist.xlsx 에서 변환)",
        "",
        "■ 채워야 하는 것",
        "  하한 / 상한 — 이 재료를 실제로 얼마부터 얼마까지 쓰는가 (%, 완제품 기준).",
        "     비어 있으면 앱이 임시 추정치를 쓰고 경고를 띄운다. 값이 채워질수록",
        "     제안 숫자를 믿을 수 있다. 이것이 이 표의 존재 이유다.",
        "  슬롯       — 기능군. 슬롯 하나가 모형의 변수 하나다.",
        "                같은 슬롯에 재료가 여럿이면 총량+비율로 묶어 쓴다.",
        "",
        "■ 자동으로 채워지는 것 (건드리지 않아도 된다)",
        "  담당축     — 온톨로지가 계산한, 이 재료가 움직이는 목표축과 방향.",
        "                비어 있으면 이 프로파일에서 아무 축도 못 움직인다는 뜻이다.",
        "",
        "■ 등급",
        "  필수 — 반드시 들어간다. 0 수준을 쓰지 않는다(DoE 규칙 3).",
        "  권장 — 기본 팔레트.",
        "  옵션 — 필요할 때.",
        "  제한 — 쓸 수 있으나 목표와 상충(저당·라벨 등). 앱이 경고한다.",
        "  제외 — 쓰지 않는다. 앱이 팔레트에서 뺀다.",
        "",
        "■ 확인 열",
        "  'ID 확인 필요'  — 온톨로지에서 그 ID 를 못 찾았다. 오탈자이거나",
        "                    아직 온톨로지에 없는 재료다. 고치거나 행을 지운다.",
        "  '슬롯·등급 확인 필요' — 실측 데이터에서 자동으로 만든 행이다.",
        "",
        "■ 고친 뒤",
        "  앱을 다시 열면 바로 반영된다. 생성기를 다시 돌리면 이 파일을 덮어쓰므로",
        "  손으로 채운 값이 사라진다 — 다시 돌릴 일이 있으면 먼저 복사해 둘 것.",
    ]:
        ws2.append([line])
    ws2.column_dimensions["A"].width = 96

    if unresolved:
        ws3 = wb.create_sheet("확인필요")
        ws3.append(["재료", "값", "문제"])
        for c in ws3[1]:
            c.font = Font(bold=True)
        for u in unresolved:
            ws3.append(list(u))
        for i, w in enumerate([28, 40, 30], 1):
            ws3.column_dimensions[get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    return records, unresolved


if __name__ == "__main__":
    recs, un = build()
    print(f"작성: {OUT}")
    print(f"  행 {len(recs)}개")
    byp = {}
    for r in recs:
        byp.setdefault(r["프로파일"], []).append(r)
    for p, rs in byp.items():
        filled = sum(1 for r in rs if r["하한"] != "" or r["상한"] != "")
        noaxis = sum(1 for r in rs if not r["담당축"] and not r["확인"])
        print(f"  [{p}] {len(rs)}행 · 범위 있음 {filled} · 담당축 없음 {noaxis}")
    if un:
        print(f"  확인 필요 {len(un)}건:")
        for n, v, why in un:
            print(f"     {n[:22]:24s} {v[:38]:40s} {why}")
