# -*- coding: utf-8 -*-
"""
DRAFT 검토표 생성 — 판단이 들어간 곳을 한 장에 모은다.

왜 필요한가
-----------
온톨로지 결손을 메우면서 여러 파일에 엣지를 추가했는데, 그중에는 기작이
자명한 것도 있고 판단이 들어간 것도 있다. 파일이 흩어져 있어 무엇을 검토해야
하는지 한눈에 안 보인다.

이 도구는 DRAFT 로 표시된 파일에서 추가된 엣지를 전부 긁어, 확신도와 note 를
붙여 xlsx 한 장으로 만든다. 승인/보류/수정 열을 두어 표에서 바로 판정할 수
있게 한다.

실행:
    python tools/review_drafts.py
    -> data/draft_review.xlsx
"""
from __future__ import annotations

import glob
import os
import sys

import yaml
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
LAYERS = os.path.join(_ROOT, "ontology_v2", "layers")
OUT = os.path.join(_ROOT, "data", "draft_review.xlsx")

COLS = ["파일", "재료/태그", "→ 대상", "방향", "크기", "확신도", "판단이 들어간 이유", "판정", "메모"]

# 확신도가 낮을수록 검토 우선순위가 높다 — 데이터가 곧 덮을 값이지만
# 그전까지는 이 숫자가 제안을 만든다.
PRIORITY = {"low": 0, "medium": 1, "high": 2}
FILL = {"low": "F6DCDA", "medium": "F6E7CE", "high": "D9EAD3"}


def draft_files():
    out = []
    for f in sorted(glob.glob(os.path.join(LAYERS, "*.yaml"))):
        head = open(f, encoding="utf-8").read(3000)
        if "DRAFT" in head:
            out.append(f)
    return out


def collect():
    rows = []
    for f in draft_files():
        name = os.path.basename(f)
        d = yaml.safe_load(open(f, encoding="utf-8")) or {}

        # C 층 — 재료/태그 확장
        for key in ("ingredients_ext", "ingredients", "function_tags_ext", "function_tags"):
            for item in (d.get(key) or []):
                owner = item.get("alias_of") or item.get("id")
                for ek in ("overrides", "effects", "flavor_profile"):
                    for e in (item.get(ek) or []):
                        rows.append(dict(
                            파일=name,
                            재료태그=str(owner).replace("ING.", "").replace("FT.", "FT:"),
                            대상=e.get("to"),
                            방향=e.get("direction"),
                            크기=e.get("magnitude") or e.get("intensity"),
                            확신도=e.get("confidence"),
                            이유=(e.get("note") or item.get("definition") or "").strip()))

        # R 층 — 관계 레코드
        for key in ("relations_proxy",):
            for r in (d.get(key) or []):
                ev = r.get("evidence") or {}
                if str(ev.get("date", "")).startswith("2026-08-2"):
                    rows.append(dict(
                        파일=name, 재료태그=r.get("id"),
                        대상=f"{r.get('parameter')} → {r.get('percept')}",
                        방향=r.get("monotone"), 크기=r.get("functional_form"),
                        확신도=r.get("confidence"),
                        이유=(r.get("note") or "").strip()))

        # M 층 — 측정 카드
        for c in (d.get("measurement_cards") or []):
            rows.append(dict(
                파일=name, 재료태그=c.get("term_id"),
                대상=f"tier={c.get('tier')} · goal={c.get('default_goal')}",
                방향="", 크기=c.get("scale_type"),
                확신도=c.get("reliability"),
                이유=(c.get("evaluation_note") or "").strip()))
        # M 층 meta 의 review_items
        meta = d.get("meta") or {}
        for it in (meta.get("review_items") or []):
            rows.append(dict(파일=name, 재료태그="(파일 전체)", 대상="review_item",
                             방향="", 크기="", 확신도="low",
                             이유=" ".join(str(it).split())))
    return rows


def main():
    rows = collect()
    rows.sort(key=lambda r: (PRIORITY.get(r.get("확신도"), 3), r["파일"], r["재료태그"]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "draft_review"
    ws.append(COLS)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E7EBE5")
    for r in rows:
        ws.append([r["파일"], r["재료태그"], r["대상"], r.get("방향") or "",
                   r.get("크기") or "", r.get("확신도") or "", r["이유"], "", ""])
    ci = COLS.index("확신도") + 1
    for i in range(2, ws.max_row + 1):
        v = ws.cell(i, ci).value
        if v in FILL:
            f = PatternFill("solid", fgColor=FILL[v])
            for c in range(1, len(COLS) + 1):
                ws.cell(i, c).fill = f
    for i, w in enumerate([34, 26, 34, 10, 9, 9, 66, 10, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws2 = wb.create_sheet("README")
    for line in [
        "DRAFT 검토표",
        "",
        "생성: tools/review_drafts.py — DRAFT 로 표시된 온톨로지 파일에서 자동 수집",
        "",
        "■ 어떻게 읽나",
        "  확신도가 낮은 것부터 정렬돼 있다. 색도 확신도다.",
        "    빨강(low)   = 추측이 많이 들어갔다. 여기부터 보시면 된다",
        "    주황(medium)= 방향은 근거가 있으나 크기가 잠정",
        "    초록(high)  = 기작이 자명하거나 문서에 근거가 있다",
        "",
        "  확신도는 모형에서 Λ 로 쓰인다. low=0.4 는 실측 한 건에도 쉽게 밀리고,",
        "  high=8 은 실험 8건어치 무게로 버틴다. 즉 확신도를 틀리게 매기면",
        "  데이터가 들어와도 사전값이 안 바뀌거나, 반대로 너무 쉽게 흔들린다.",
        "",
        "■ 판정 열에 적을 것",
        "  OK      그대로 간다",
        "  수정    메모 열에 고칠 값을 적는다 (예: 크기 strong->medium)",
        "  삭제    이 엣지를 빼야 한다",
        "  보류    실측을 보고 정한다",
        "",
        "■ 판정 후",
        "  이 파일을 그대로 두시면 제가 읽어서 반영합니다.",
        "  '삭제' 가 있으면 도달성(불변식 6)이 깨질 수 있으니 반영 후 검사를 돌립니다.",
        "",
        "■ 특히 봐주실 것",
        "  1. 쌀 우유 M 카드의 'oil flavor' 매핑 — L.tx.oiliness(식감)로 걸었다.",
        "     Layer L 에 일반 식물유 '향' 용어가 없어서다.",
        "  2. 같은 카드의 산화취 — 신선 시료 기준(sample)으로 뒀다. 기본 프로파일은",
        "     숙성 기준(sample_aged)으로 격리하는데, 실제 패널은 신선 시료로 평가했다.",
        "  3. 신선한 홍고추 향 — L 에 용어가 없어 fresh_chili_green 에 약하게 걸었다.",
        "     용어를 추가하는 편이 옳을 수 있다.",
    ]:
        ws2.append([line])
    ws2.column_dimensions["A"].width = 100

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)

    import collections
    c = collections.Counter(r.get("확신도") for r in rows)
    print(f"작성: {OUT}")
    print(f"  검토 항목 {len(rows)}건 · 확신도 분포 {dict(c)}")
    print(f"  대상 파일 {len(draft_files())}개:")
    for f in draft_files():
        print(f"     {os.path.basename(f)}")


if __name__ == "__main__":
    main()
