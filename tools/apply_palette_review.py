# -*- coding: utf-8 -*-
"""
팔레트 검토 결과를 palette.xlsx 에 반영한다.

data/palette_review.xlsx 의 판정 열을 읽어 적용한다.

  OK    아무것도 하지 않는다
  수정   '고칠 하한' / '고칠 상한' 에 값이 있으면 그 값으로. 비어 있으면 유지
  삭제   등급을 '제외' 로 바꾼다 (ID 는 남긴다 — 불변식 3)
  보류   범위근거에 보류 사유를 적고, 권장 등급이면 옵션으로 낮춘다
         (기본 팔레트에서 빠지되 필요하면 여전히 고를 수 있게)

메모 열은 범위근거 뒤에 붙여 왜 그 값이 됐는지 남긴다.

실행:
    python tools/apply_palette_review.py            # 미리보기
    python tools/apply_palette_review.py --write    # 실제 반영
"""
from __future__ import annotations

import os
import sys
from datetime import date

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
sys.path.insert(0, os.path.join(_ROOT, "app"))
sys.path.insert(0, os.path.join(_ROOT, "engine"))

import palette  # noqa: E402

REVIEW = os.path.join(_ROOT, "data", "palette_review.xlsx")
TODAY = date.today().isoformat()


def read_review():
    wb = openpyxl.load_workbook(REVIEW, data_only=True)
    ws = wb["palette_review"]
    hdr = [c.value for c in ws[1]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr, r))
        if not d.get("온톨로지ID"):
            continue
        j = str(d.get("판정") or "").strip()
        if not j:
            continue
        out.append(dict(
            profile=str(d["프로파일"]).strip(),
            gid=str(d["온톨로지ID"]).strip(),
            verdict=j.upper() if j.upper() == "OK" else j,
            lo=d.get("고칠 하한"), hi=d.get("고칠 상한"),
            memo=str(d.get("메모") or "").strip()))
    return out


def _num(v):
    if v is None or (isinstance(v, str) and not str(v).strip()):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main(write=False):
    review = read_review()
    by = {}
    for r in review:
        by[(r["profile"], r["gid"])] = r

    changes = []
    for prof in palette.profiles_in():
        t = palette.load(prof)
        rows = []
        for row in t.rows:
            d = dict(row)
            key = (prof, row["온톨로지ID"])
            r = by.get(key)
            if not r or r["verdict"].upper() == "OK":
                rows.append(d)
                continue

            v = r["verdict"]
            before = (row["하한"], row["상한"], row["등급"])

            if v == "삭제":
                d["등급"] = "제외"
                d["범위근거"] = f"[검토 {TODAY} · 삭제] {r['memo'] or '팔레트에서 뺀다'}"
            elif v == "보류":
                if d["등급"] == "권장":
                    d["등급"] = "옵션"      # 기본 팔레트에서만 빼고 남겨 둔다
                d["범위근거"] = f"[검토 {TODAY} · 보류] {r['memo']} · 이전: {row['범위근거']}"
            elif v == "수정":
                lo, hi = _num(r["lo"]), _num(r["hi"])
                if lo is not None:
                    d["하한"] = lo
                if hi is not None:
                    d["상한"] = hi
                tag = "값 수정" if (lo is not None or hi is not None) else "메모만"
                d["범위근거"] = f"[검토 {TODAY} · {tag}] {r['memo']} · 이전: {row['범위근거']}"
            else:
                rows.append(d)
                continue

            after = (d["하한"], d["상한"], d["등급"])
            if before != after or v in ("수정", "보류", "삭제"):
                changes.append((prof, row["온톨로지ID"], v, before, after, r["memo"][:50]))
            rows.append(d)

        if write:
            palette.save_rows(prof, rows)

    print(f"{'반영' if write else '미리보기'} — 변경 {len(changes)}건\n")
    for prof, gid, v, b, a, memo in changes:
        g = gid.replace("ING.", "")
        if b[:2] != a[:2]:
            print(f"  [{v}] {g:28s} {b[0]}~{b[1]} -> {a[0]}~{a[1]}")
        elif b[2] != a[2]:
            print(f"  [{v}] {g:28s} 등급 {b[2]} -> {a[2]}")
        else:
            print(f"  [{v}] {g:28s} (메모만)")
        if memo:
            print(f"        {memo}")

    if write:
        print()
        for prof in palette.profiles_in():
            t = palette.load(prof)
            s = t.status()
            bad = t.bad_ranges()
            print(f"  [{prof:22s}] 쓸 수 있는 재료 {s['usable']} · 범위 {s['with_range']}"
                  + (f" · !! 뒤집힌 범위 {bad}" if bad else ""))
    else:
        print("\n실제로 반영하려면: python tools/apply_palette_review.py --write")


if __name__ == "__main__":
    main(write="--write" in sys.argv)
