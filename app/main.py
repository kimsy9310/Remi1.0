# -*- coding: utf-8 -*-
"""
Remi 1.0 — 웜루프 앱

    실측 데이터 → 사전값 보정 → 다음 실험 제안 → (랩) → 다시 실측

실행:  streamlit run app/main.py
"""
from __future__ import annotations

import os
import sys

import numpy as np
import pandas as pd
import streamlit as st

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
for p in (os.path.join(_ROOT, "engine"), _HERE):
    if p not in sys.path:
        sys.path.insert(0, p)

import dataio                                    # noqa: E402
import palette as pal                            # noqa: E402
import store                                     # noqa: E402
import warmloop as wl                            # noqa: E402
from formulator.v2adapter import V2Ontology      # noqa: E402

try:                                    # 제안 근거 표시에만 쓴다
    import scipy  # noqa: F401
    _HAS_SCIPY = True
except ImportError:
    _HAS_SCIPY = False

st.set_page_config(page_title="Remi 1.0", page_icon="🧪", layout="wide")

DATA_DIR = os.path.join(_ROOT, "data")


# ---------------------------------------------------------------- 캐시
@st.cache_resource(show_spinner="온톨로지 로드 중…")
def get_onto():
    return V2Ontology()


@st.cache_data(show_spinner="실측 데이터 읽는 중…")
def get_data(path, profile, _mtime):
    return dataio.load_warmloop(path, get_onto(), profile)


@st.cache_data(show_spinner=False)
def get_palette(profile, _mtime, variant=None):
    return pal.load(profile, variant=variant)


def variants_of(profile):
    """이 프로파일에 정의된 목적 목록. 기본(빈 값)은 항상 첫째."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(pal.DEFAULT_PATH, data_only=True)
        ws = wb["palette"]
        hdr = [str(c.value).strip() if c.value else "" for c in ws[1]]
        iP, iV = hdr.index("프로파일"), hdr.index("목적")
        out = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if str(r[iP] or "").strip() != profile:
                continue
            v = str(r[iV] or "").strip()
            if v and v not in out:
                out.append(v)
        return out
    except Exception:                                          # noqa: BLE001
        return []


def load_palette(profile, variant=None):
    """팔레트 표. 없거나 이 프로파일 행이 없으면 None 을 준다(앱은 계속 돈다)."""
    try:
        t = get_palette(profile, os.path.getmtime(pal.DEFAULT_PATH), variant)
        return t if t.rows else None
    except FileNotFoundError:
        return None
    except Exception as e:                                        # noqa: BLE001
        st.sidebar.error(f"팔레트 표를 읽지 못했습니다: {e}")
        return None


def core_terms(profile):
    cards = onto._ref.load_cards(profile, onto.layers)
    return [c["term_id"] for c in cards
            if c["tier"] == "core" and c["evidence_required"] != "sample_aged"]


def axis_label(t):
    return t.split(".")[-1]


def ing_label(g):
    return g.replace("ING.", "")


# ---------------------------------------------------------------- 사이드바
onto = get_onto()

st.sidebar.title("Remi 1.0")
st.sidebar.caption("식품 레시피 포뮬레이터 · 웜루프")

profiles = sorted(onto.profiles)
default_ix = profiles.index("beverage_rice_milk") if "beverage_rice_milk" in profiles else 0
profile = st.sidebar.selectbox("프로파일", profiles, index=default_ix)

xlsx = sorted(f for f in os.listdir(DATA_DIR) if f.endswith(".xlsx") and not f.startswith("~"))
if not xlsx:
    st.sidebar.error("data/ 에 xlsx 가 없습니다.")
    st.stop()
fname = st.sidebar.selectbox("실측 데이터", xlsx)
fpath = os.path.join(DATA_DIR, fname)

_vars = variants_of(profile)
variant = None
if _vars:
    _pick = st.sidebar.selectbox("목적", ["(기본)"] + _vars,
                                 help="같은 제형이라도 목적에 따라 재료 범위가 달라집니다.")
    variant = None if _pick == "(기본)" else _pick

st.sidebar.divider()
st.sidebar.caption(
    f"온톨로지 재료 {len(onto.ingredients)}종 · 태그 {len(onto.tags)}종\n\n"
    f"필러 `{onto.filler_of(profile)}`")

# ---------------------------------------------------------------- 데이터 로드
try:
    data = get_data(fpath, profile, os.path.getmtime(fpath))
except Exception as e:                                            # noqa: BLE001
    st.error(f"**데이터를 읽지 못했습니다**\n\n{e}")
    st.info(
        "이 프로파일의 M 카드에 `legacy_column` 이 있어야 관능 컬럼이 L.* 축과 이어집니다. "
        "`ontology_v2/layers/layerM_cards_*.yaml` 을 확인하세요.")
    st.stop()

tab_data, tab_learn, tab_suggest, tab_bounds, tab_entry = st.tabs(
    ["① 데이터", "② 학습", "③ 제안", "④ 팔레트", "⑤ 실험 입력"])

# ================================================================= ① 데이터
with tab_data:
    st.subheader("무엇이 들어왔나")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("샘플", len(data.sample_ids))
    c2.metric("재료", len(data.names))
    c3.metric("관능 축", len(data.y_terms))
    c4.metric("벤치마크", len(data.benchmark_ids) or "—")

    for n in data.notes:
        st.warning(n, icon="⚠️")
    if data.unmapped_columns:
        st.warning(f"온톨로지 ID 로 잇지 못한 컬럼: {data.unmapped_columns}", icon="⚠️")
    if data.missing_axes:
        st.info(f"카드는 있으나 데이터에 컬럼이 없는 축: "
                f"{[axis_label(t) for t in data.missing_axes]}")

    st.markdown("##### 관능 (벤치마크 상대, −3…+3)")
    st.dataframe(
        {"sample": data.sample_ids,
         **{axis_label(t): data.Y[:, k].tolist() for k, t in enumerate(data.y_terms)}},
        use_container_width=True, hide_index=True)

    with st.expander("배합 (합계 100으로 정규화)"):
        st.caption(f"원본 배치 합계: {np.round(data.raw_totals, 1).tolist()}")
        st.dataframe(
            {"sample": data.sample_ids,
             **{ing_label(g): np.round(data.X[:, i], 3).tolist()
                for i, g in enumerate(data.names)}},
            use_container_width=True, hide_index=True)

# ================================================================= ② 학습
with tab_learn:
    st.subheader("사전값을 실측이 얼마나 밀어냈나")
    ptab = load_palette(profile, variant)
    bounds = ptab.bounds() if ptab else {}
    try:
        res = wl.run(onto, data, profile, bounds=bounds or None)
    except Exception as e:                                        # noqa: BLE001
        st.error(f"적합에 실패했습니다: {e}")
        st.stop()

    c1, c2, c3 = st.columns(3)
    c1.metric("적합에 쓴 샘플", res.n)
    c2.metric("탐색된 독립 방향", res.eff_directions,
              delta=f"자유변수 {len(res.free)}개 중",
              delta_color="off")
    c3.metric("부호가 뒤집힌 칸", len(res.contradicted()))

    if res.eff_directions < len(res.free):
        st.info(
            f"샘플은 {res.n}건이지만 실제로 탐색된 독립 방향은 **{res.eff_directions}개**입니다. "
            f"재료를 함께 움직이면 여러 런이 같은 방향을 반복합니다 — "
            f"개수보다 정보량이 중요한 이유입니다.", icon="📐")
    for w in res.warnings:
        st.warning(w, icon="⚠️")

    st.markdown("##### 가장 많이 움직인 칸")
    sh = res.shifts(20)
    if sh:
        st.dataframe(
            {"재료": [ing_label(s["ingredient"]) for s in sh],
             "축": [axis_label(s["axis"]) for s in sh],
             "사전": [round(s["prior"], 2) for s in sh],
             "실측 보정": [round(s["post"], 2) for s in sh],
             "이동": [round(s["shift"], 2) for s in sh],
             "Λ(확신도)": [s["lam"] for s in sh],
             "부호뒤집힘": ["예" if s["flipped"] else "" for s in sh]},
            use_container_width=True, hide_index=True)

    cc = res.contradicted()
    if cc:
        st.markdown("##### 온톨로지와 부호가 반대인 칸")
        st.caption(
            "사전지식이 틀렸을 수도, 실험이 교락됐을 수도 있습니다. Λ 가 낮고 이동이 작으면 "
            "약한 증거이니 성급히 온톨로지를 고치지 마세요.")
        for s in cc:
            st.write(
                f"- **{ing_label(s['ingredient'])}** → *{axis_label(s['axis'])}* : "
                f"사전 `{s['prior']:+.2f}` / 실측 `{s['post']:+.2f}`  (Λ={s['lam']})")

    ls = res.learned_from_silence()
    if ls:
        with st.expander(f"사전이 비어 있던 칸을 데이터가 채운 것 ({len(ls)}개)"):
            st.caption("온톨로지에 엣지가 없던 자리입니다. 반복해서 확인되면 온톨로지에 "
                       "추가할 후보입니다.")
            st.dataframe(
                {"재료": [ing_label(s["ingredient"]) for s in ls],
                 "축": [axis_label(s["axis"]) for s in ls],
                 "실측 계수": [round(s["post"], 2) for s in ls]},
                use_container_width=True, hide_index=True)

    st.session_state["res"] = res

# ================================================================= ③ 제안
with tab_suggest:
    st.subheader("다음 실험 제안")
    res = st.session_state.get("res")
    if res is None:
        st.info("먼저 ② 학습 탭을 여세요.")
        st.stop()

    ptab = load_palette(profile, variant)
    bounds = ptab.bounds() if ptab else {}
    if not bounds:
        st.warning(
            "재료 범위가 하나도 없습니다. 상한 없이 제안하면 물리적으로 불가능한 배합이 "
            "나올 수 있습니다. **④ 팔레트** 탭에서 먼저 넣어주세요.", icon="🚧")
    elif ptab:
        miss = [g for g in res.free if g not in bounds]
        if miss:
            st.info(
                f"범위가 없는 재료 {len(miss)}종은 경계 없이 움직입니다: "
                f"{[ing_label(g) for g in miss[:8]]}", icon="ℹ️")

    base_ix = st.selectbox(
        "기준 배합", range(len(res_ids := data.sample_ids)),
        format_func=lambda i: res_ids[i],
        index=len(res_ids) - 1,
        help="여기서 출발해 목표 쪽으로 움직입니다.")
    order = [data.names.index(g) for g in res.names]
    x0 = data.X[base_ix][order]

    y_base = res.model.predict(x0[None, :])[0]

    st.markdown("##### 목표 (벤치마크 대비)")
    st.caption(
        "척도가 벤치마크 상대라 **0 은 '벤치마크와 같게'** 라는 뜻입니다 — '현재 유지' 가 "
        "아닙니다. 그래서 슬라이더 기본값을 **기준 배합의 현재 예측값**으로 둡니다. "
        "한 축만 움직이면 그 요청만 반영됩니다. "
        "신경 쓰지 않는 축은 **자유**로 두세요. 자유가 아니면 그 축도 목표로 붙잡으므로, "
        "정작 조절하려는 축이 눌립니다.")

    cols = st.columns(len(res.y_terms))
    target, free_axes = {}, []
    for c, t, yb in zip(cols, res.y_terms, y_base):
        c.markdown(f"**{axis_label(t)}**")
        fr = c.checkbox("자유", key=f"fr_{t}",
                        help="이 축을 목적함수에서 뺍니다. 결과가 어떻게 되든 상관없을 때.")
        # 스텝을 0.05 로 두어 기본값이 현재 예측값과 정확히 맞게 한다.
        # 0.25 스텝이면 기본값이 현재값에서 최대 0.12 어긋나고, 손대지 않아도
        # 그만큼이 '요청' 으로 잡혀 배합이 움직인다 — 무엇이 내 조작인지 흐려진다.
        v = c.slider(" ", -3.0, 3.0, float(np.clip(round(yb, 2), -3, 3)), 0.05,
                     key=f"tg_{t}", label_visibility="collapsed", disabled=fr)
        c.caption(f"현재 {yb:+.2f}")
        if fr:
            free_axes.append(t)
        target[t] = v

    if len(free_axes) == len(res.y_terms):
        st.warning("모든 축을 자유로 두면 최적화할 목표가 없습니다. 하나 이상 풀어주세요.",
                   icon="⚠️")
        st.stop()

    # ---- 슬라이더가 바뀌면 즉시 재계산 (버튼 없음)
    try:
        x = wl.suggest(res, target, x0, bounds=bounds or None,
                       free_axes=free_axes or None)
    except Exception as e:                                        # noqa: BLE001
        st.error(f"제안에 실패했습니다: {e}")
        st.stop()

    delta = x - x0
    y0 = res.model.predict(x0[None, :])[0]
    y1 = res.model.predict(x[None, :])[0]
    moved = np.abs(delta) > 1e-3
    # '요청' 은 현재 예측과 다르게 잡은 축이다. 슬라이더 기본값이 현재값이므로
    # 움직이지 않은 축은 요청이 아니다.
    # 한 스텝(0.05)의 절반보다 작게 움직인 것은 반올림 잔차이지 요청이 아니다.
    tgt_axes = [t for k, t in enumerate(res.y_terms)
                if t not in free_axes and abs(target[t] - y_base[k]) > 0.026]

    # ---- 목표를 아무것도 안 줬을 때
    if not tgt_axes:
        st.info("슬라이더가 모두 현재값 그대로입니다. 하나를 움직여 보세요 — "
                "그 축만 요청으로 잡히고 나머지는 지금 수준에서 유지됩니다.", icon="🎚️")

    # ---- 왜 안 움직이는가 진단 (이게 없으면 '반응이 없다'로만 보인다)
    elif not moved.any():
        why = []
        for t in tgt_axes:
            k = res.y_terms.index(t)
            gap = target[t] - y0[k]
            if abs(gap) < 0.02:
                why.append(f"**{axis_label(t)}** 는 기준 배합이 이미 목표에 있습니다"
                           f"(현재 {y0[k]:+.2f}).")
                continue
            movers = [g for i, g in enumerate(res.free)
                      if abs(res.Gamma_post[i, k]) > 1e-9]
            if not movers:
                why.append(f"**{axis_label(t)}** 를 움직일 수 있는 재료가 이 팔레트에 "
                           f"없습니다. 온톨로지에 엣지가 없거나 팔레트에서 빠졌습니다.")
                continue
            pinned = []
            for g in movers:
                if g in bounds:
                    v = x[res.names.index(g)]
                    lo, hi = bounds[g]
                    if v <= lo + 1e-6:
                        pinned.append(f"{ing_label(g)}(하한 {lo})")
                    elif v >= hi - 1e-6:
                        pinned.append(f"{ing_label(g)}(상한 {hi})")
            if pinned and len(pinned) == len(movers):
                why.append(f"**{axis_label(t)}** 를 움직일 재료가 전부 경계에 걸려 "
                           f"있습니다: {', '.join(pinned)}. ④ 팔레트에서 범위를 넓히세요.")
            else:
                why.append(f"**{axis_label(t)}** 는 다른 축을 유지하라는 제약과 "
                           f"상충해 움직이지 못했습니다. 유지할 축을 줄여 보세요.")
        st.warning("배합이 바뀌지 않았습니다. 이유:\n\n" + "\n\n".join(f"- {w}" for w in why),
                   icon="🔎")

    # ---- 배합 변경: 바뀐 것만, 큰 순서로, 색으로
    st.markdown("##### 배합 변경")
    if moved.any():
        idx = np.argsort(-np.abs(delta))
        rows = []
        for i in idx:
            if not moved[i]:
                continue
            g = res.names[i]
            lo, hi = bounds.get(g, (None, None))
            pin = ""
            if lo is not None:
                if x[i] <= lo + 1e-6:
                    pin = f"하한 {lo:g}"
                elif x[i] >= hi - 1e-6:
                    pin = f"상한 {hi:g}"
            rows.append({
                "재료": ing_label(g),
                "기준": round(float(x0[i]), 3),
                "제안": round(float(x[i]), 3),
                "변화": round(float(delta[i]), 3),
                "": "▲" if delta[i] > 0 else "▼",
                "경계": pin,
            })
        df = pd.DataFrame(rows)
        vmax = float(np.abs(delta).max()) or 1.0

        def _tint(v):
            """증가는 초록, 감소는 빨강. 진하기는 변화 크기에 비례.
            matplotlib 없이 인라인 CSS 로 칠한다 — 의존성을 늘리지 않으려고."""
            try:
                f = float(v)
            except (TypeError, ValueError):
                return ""
            if abs(f) < 1e-9:
                return ""
            a = 0.15 + 0.45 * min(abs(f) / vmax, 1.0)
            rgb = "46,139,87" if f > 0 else "178,34,34"
            return f"background-color: rgba({rgb},{a:.2f})"

        st.dataframe(
            df.style
              .map(_tint, subset=["변화"])
              .format({"기준": "{:.3f}", "제안": "{:.3f}", "변화": "{:+.3f}"}),
            use_container_width=True, hide_index=True)
        n_un = int((~moved).sum())
        if n_un:
            with st.expander(f"변화 없는 재료 {n_un}종"):
                st.dataframe(
                    {"재료": [ing_label(g) for i, g in enumerate(res.names) if not moved[i]],
                     "값": [round(float(x0[i]), 3) for i in range(len(x0)) if not moved[i]]},
                    use_container_width=True, hide_index=True)
    else:
        st.caption("바뀐 재료가 없습니다.")
    st.caption(f"합계 {x.sum():.4f}")

    # ---- 예측 관능: 목표 · 기준 · 제안 · 달성
    st.markdown("##### 예측 관능")
    ach = []
    for k, t in enumerate(res.y_terms):
        gap0, gap1 = abs(target[t] - y0[k]), abs(target[t] - y1[k])
        ach.append("—" if gap0 < 1e-9 else ("달성" if gap1 < 0.05 else
                   f"{max(0.0, (gap0 - gap1) / gap0) * 100:.0f}%"))
    st.dataframe(
        {"축": [axis_label(t) for t in res.y_terms],
         "목표": [target[t] for t in res.y_terms],
         "기준 예측": np.round(y0, 2).tolist(),
         "제안 예측": np.round(y1, 2).tolist(),
         "목표 접근": ach},
        use_container_width=True, hide_index=True)

    if np.abs(y1).max() > 3.0:
        st.error(
            f"예측이 ±3 관능 척도를 벗어났습니다(최대 {np.abs(y1).max():.1f}). "
            f"재료 범위가 없거나 너무 넓을 때 생깁니다.", icon="🚨")

    # ---- 근거: 모형이 무엇을 보고 이렇게 골랐나
    with st.expander("이 제안이 나온 근거"):
        if not tgt_axes:
            st.caption("목표를 주면 여기에 근거가 나옵니다.")
        else:
            st.markdown("**목표로 준 축을 움직이는 재료와 그 계수**")
            st.caption(
                "계수는 '이 재료를 1 표준편차 올릴 때 그 축이 몇 칸 움직이나'입니다. "
                "실측이 있으면 보정된 값이고, 없으면 온톨로지 사전값입니다. "
                "부호가 목표와 같아야 그 재료를 올립니다.")
            for t in tgt_axes:
                k = res.y_terms.index(t)
                items = [(res.free[i], float(res.Gamma_post[i, k]))
                         for i in range(len(res.free))
                         if abs(res.Gamma_post[i, k]) > 1e-9]
                items.sort(key=lambda z: -abs(z[1]))
                st.markdown(f"**{axis_label(t)}**  (목표 {target[t]:+.2f})")
                if not items:
                    st.caption("   이 축을 움직이는 재료가 팔레트에 없습니다.")
                    continue
                st.dataframe(
                    {"재료": [ing_label(g) for g, _ in items[:10]],
                     "계수": [round(v, 3) for _, v in items[:10]],
                     "이번 변화": [round(float(delta[res.names.index(g)]), 3)
                                for g, _ in items[:10]]},
                    use_container_width=True, hide_index=True)
            st.divider()
            st.markdown("**경계에 걸린 재료**")
            pinned = []
            for i, g in enumerate(res.names):
                if g in bounds:
                    lo, hi = bounds[g]
                    if x[i] <= lo + 1e-6:
                        pinned.append(f"{ing_label(g)} = 하한 {lo:g}")
                    elif x[i] >= hi - 1e-6:
                        pinned.append(f"{ing_label(g)} = 상한 {hi:g}")
            st.caption(", ".join(pinned) if pinned else "없음 — 모두 범위 안쪽입니다.")
            st.caption(
                f"풀이: {'scipy SLSQP' if _HAS_SCIPY else '경사하강(scipy 없음)'} · "
                f"자유변수 {res.model.q} · 필러 {ing_label(res.filler)}")

    # ---- 이력 저장은 명시적으로
    if st.button("이 제안을 이력에 저장", type="primary"):
        store.log_run(profile, fname, res.n, res.y_terms,
                      target={axis_label(k): v for k, v in target.items()},
                      proposal={ing_label(g): round(float(v), 4)
                                for g, v in zip(res.names, x)})
        st.success("저장했습니다.")

    hist = store.recent_runs(5)
    if hist:
        with st.expander("최근 제안 이력"):
            for r in hist:
                st.write(f"`{r[1]}` · {r[2]} · 샘플 {r[3]}건 · 목표 {r[4]}")


# ================================================================= ④ 팔레트
with tab_bounds:
    st.subheader("팔레트와 작업 범위")
    ptab = load_palette(profile, variant)

    if ptab is None:
        st.error(
            "이 프로파일의 팔레트 행이 없습니다.\n\n"
            "`python tools/build_palette.py` 로 표를 만들거나, "
            "`data/palette.xlsx` 에 이 프로파일 행을 추가하세요.")
        st.stop()

    stt = ptab.status()
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("쓸 수 있는 재료", stt["usable"])
    c2.metric("범위 입력됨", stt["with_range"],
              delta=f"-{stt['missing']}종 남음" if stt["missing"] else "완료",
              delta_color="inverse" if stt["missing"] else "normal")
    c3.metric("제외 등급", stt["excluded"])
    c4.metric("확인 필요", stt["flagged"])

    st.progress(stt["with_range"] / max(stt["usable"], 1),
                text=f"범위 {stt['with_range']} / {stt['usable']}")
    st.caption(
        "온톨로지의 `limitations` 는 자유 서술이라 기계가 읽지 못합니다(스펙 G5). "
        "여기 값이 제안의 상·하한이자 사전계수의 스케일이 됩니다. "
        "**채워질수록 제안 숫자를 믿을 수 있습니다.** 원본은 `data/palette.xlsx`.")

    # ---- 목표축 커버리지 (G2)
    core = core_terms(profile)
    cov = ptab.coverage(core)
    dead = [t for t, v in cov.items() if not v]
    st.markdown("##### 목표축 커버리지")
    st.dataframe(
        {"목표축": [axis_label(t) for t in core],
         "이 축을 움직이는 재료": [len(cov[t]) for t in core],
         "예": [", ".join(ing_label(g) for g in cov[t][:4]) or "— 없음" for t in core]},
        use_container_width=True, hide_index=True)
    if dead:
        st.error(
            f"움직일 수 있는 재료가 없는 축: **{[axis_label(t) for t in dead]}** — "
            f"목표로 줘도 반응하지 않습니다. 해당 축을 움직이는 재료를 팔레트에 넣거나, "
            f"온톨로지에 엣지가 없는 경우입니다.", icon="🚨")

    na = ptab.no_axis()
    if na:
        st.warning(
            f"이 프로파일에서 아무 목표축도 못 움직이는 재료 {len(na)}종: "
            f"{[ing_label(g) for g in na]} — 배합에는 넣을 수 있지만 모형은 조종 수단으로 "
            f"쓰지 못합니다.", icon="⚠️")
    fl = ptab.flagged()
    if fl:
        with st.expander(f"확인 필요 {len(fl)}건"):
            for g, why in fl:
                st.write(f"- `{ing_label(g)}` — {why}")
    bad = ptab.bad_ranges()
    if bad:
        st.error(f"상한이 하한보다 작거나 같은 행: {bad}", icon="🚨")

    # ---- 편집
    st.markdown("##### 표 편집")
    st.caption("하한·상한을 채우고 저장하면 `data/palette.xlsx` 에 바로 반영됩니다.")

    order = {g: i for i, g in enumerate(["필수", "권장", "옵션", "제한", "제외"])}
    rows = sorted(ptab.rows,
                  key=lambda r: (order.get(r["등급"], 9), r["슬롯"], r["온톨로지ID"]))
    edited = st.data_editor(
        [{"슬롯": r["슬롯"], "재료": r["재료"], "온톨로지ID": r["온톨로지ID"],
          "등급": r["등급"], "하한": r["하한"], "상한": r["상한"],
          "범위근거": r["범위근거"], "담당축": r["담당축"],
          "메모": r["메모"], "확인": r["확인"]} for r in rows],
        use_container_width=True, hide_index=True, num_rows="fixed",
        column_config={
            "등급": st.column_config.SelectboxColumn(
                options=["필수", "권장", "옵션", "제한", "제외"], width="small"),
            "하한": st.column_config.NumberColumn(min_value=0.0, step=0.05, format="%.3f"),
            "상한": st.column_config.NumberColumn(min_value=0.0, step=0.05, format="%.3f"),
            "온톨로지ID": st.column_config.TextColumn(disabled=True),
            "담당축": st.column_config.TextColumn(disabled=True, help="온톨로지가 계산합니다"),
        },
        key=f"pal_{profile}")

    if st.button("저장", type="primary"):
        out, bad2 = [], []
        for e in edited:
            lo, hi = e.get("하한"), e.get("상한")
            if lo is not None and hi is not None and hi <= lo:
                bad2.append(f"{ing_label(e['온톨로지ID'])}: 상한({hi}) ≤ 하한({lo})")
                continue
            out.append({"프로파일": profile, "목적": variant or "",
                        "슬롯": e["슬롯"], "재료": e["재료"],
                        "온톨로지ID": e["온톨로지ID"], "등급": e["등급"],
                        "하한": lo, "상한": hi, "범위근거": e["범위근거"],
                        "담당축": e["담당축"], "메모": e["메모"], "확인": e["확인"]})
        if bad2:
            for b in bad2:
                st.error(b)
        else:
            pal.save_rows(profile, out, variant=variant)
            st.cache_data.clear()
            st.success(f"{len(out)}행을 저장했습니다.")
            st.rerun()

    with st.expander("슬롯 구성"):
        st.caption(
            "슬롯 하나가 모형의 변수 하나입니다. 같은 슬롯에 재료가 여럿이면 "
            "각각을 독립 변수로 두지 말고 총량+비율로 묶는 편이 낫습니다 — "
            "죽은 코너가 사라지고 런이 낭비되지 않습니다.")
        for slot, gs in ptab.slots().items():
            st.write(f"**{slot}** ({len(gs)}) — {', '.join(ing_label(g) for g in gs)}")


# ================================================================= ⑤ 실험 입력
with tab_entry:
    st.subheader("실험 데이터 넣기")
    st.caption(
        "새 배치를 시작하려면 **① 양식 만들기** 로 빈 파일을 받아 랩에 넘기고, "
        "평가가 끝나면 **② 올리기** 로 되돌려 놓으세요. 올린 파일은 `data/` 에 "
        "저장되어 사이드바에서 고를 수 있게 됩니다.")

    ptab_e = load_palette(profile, variant)

    c1, c2 = st.columns(2)

    # ---------------------------------------------------------- ① 양식 만들기
    with c1:
        st.markdown("##### ① 양식 만들기")
        if ptab_e is None:
            st.warning("팔레트 표가 없어 재료 목록을 만들 수 없습니다.", icon="🚧")
        else:
            grades = st.multiselect(
                "넣을 재료 등급", ["필수", "권장", "옵션", "제한"],
                default=["필수", "권장"],
                help="팔레트에서 이 등급인 재료만 배합 열로 만듭니다.")
            n_s = st.number_input("샘플 행 수", 1, 40, 5, 1)
            prefix = st.text_input("샘플 ID 접두사", value=profile.split("_")[0].upper()[:3],
                                   help="예: RM -> RM01, RM02 …")
            pal_e = [r["온톨로지ID"] for r in ptab_e.rows if r["등급"] in grades]
            nb = [g for g in pal_e if g not in ptab_e.bounds()]
            st.caption(f"재료 {len(pal_e)}종"
                       + (f" · 범위 없는 재료 {len(nb)}종" if nb else " · 범위 모두 있음"))

            if st.button("양식 만들기", type="primary", disabled=not pal_e):
                filler_e = onto.filler_of(profile)
                plist = list(pal_e) + ([filler_e] if filler_e not in pal_e else [])
                sids = [f"{prefix}{i:02d}" for i in range(1, int(n_s) + 1)]
                tag = f"_{variant}" if variant else ""
                out = os.path.join(DATA_DIR, f"{profile}{tag}_batch.xlsx")
                if os.path.exists(out):
                    seq = 2
                    while os.path.exists(
                            os.path.join(DATA_DIR, f"{profile}{tag}_batch{seq}.xlsx")):
                        seq += 1
                    out = os.path.join(DATA_DIR, f"{profile}{tag}_batch{seq}.xlsx")
                dataio.write_template(out, onto, profile, plist, sample_ids=sids)
                st.success(f"만들었습니다: `{os.path.basename(out)}`")
                with open(out, "rb") as fh:
                    st.download_button("내려받기", fh.read(),
                                       file_name=os.path.basename(out),
                                       mime="application/vnd.openxmlformats-officedocument."
                                            "spreadsheetml.sheet")
                st.caption(
                    "recipes 시트에 배합을, sensory 시트에 벤치마크 상대 점수(−3…+3)를 "
                    "채우세요. 벤치마크 자신은 전 축 0 으로 적고 `is_benchmark=yes` 로 "
                    "표시합니다. 축 정의는 README 시트에 있습니다.")

    # ---------------------------------------------------------- ② 올리기
    with c2:
        st.markdown("##### ② 채운 파일 올리기")
        up = st.file_uploader("xlsx 파일", type=["xlsx"], key="up_data")
        if up is not None:
            tmp = os.path.join(DATA_DIR, f"_검사중_{up.name}")
            with open(tmp, "wb") as fh:
                fh.write(up.getbuffer())
            try:
                chk = dataio.load_warmloop(tmp, onto, profile)
            except Exception as e:                             # noqa: BLE001
                os.remove(tmp)
                st.error(f"읽지 못했습니다: {e}")
                st.info(
                    "컬럼 이름이 양식과 같은지 확인해 주세요. 관능 컬럼은 M 카드의 "
                    "`legacy_column` 또는 축 ID 와 정확히 일치해야 합니다.")
            else:
                st.success(f"읽었습니다 — 샘플 {len(chk.sample_ids)}건 · "
                           f"재료 {len(chk.names)}종 · 축 {len(chk.y_terms)}개")
                for n in chk.notes:
                    st.warning(n, icon="⚠️")
                if chk.unmapped_columns:
                    st.warning(f"온톨로지 ID 로 잇지 못한 컬럼: {chk.unmapped_columns}",
                               icon="⚠️")
                st.dataframe(
                    {"sample": chk.sample_ids,
                     **{axis_label(t): chk.Y[:, k].tolist()
                        for k, t in enumerate(chk.y_terms)}},
                    use_container_width=True, hide_index=True)

                dest = os.path.join(DATA_DIR, up.name)
                exists = os.path.exists(dest)
                if exists:
                    st.warning(f"`{up.name}` 이 이미 있습니다. 저장하면 덮어씁니다.",
                               icon="⚠️")
                if st.button("이 데이터로 저장", type="primary"):
                    os.replace(tmp, dest)
                    st.cache_data.clear()
                    st.success(f"저장했습니다: `{up.name}` — "
                               f"사이드바 '실측 데이터' 에서 고르면 학습에 들어갑니다.")
                    st.rerun()
                else:
                    st.caption("저장을 누르기 전까지는 반영되지 않습니다.")

    # ---------------------------------------------------------- 현재 데이터
    st.divider()
    st.markdown("##### `data/` 에 있는 실측 파일")
    rows_e = []
    for f in sorted(os.listdir(DATA_DIR)):
        if not f.endswith(".xlsx") or f.startswith(("~", "_검사중_")):
            continue
        if f in ("palette.xlsx", "draft_review.xlsx", "palette_review.xlsx"):
            continue
        fp = os.path.join(DATA_DIR, f)
        got = "—"
        for pf in sorted(onto.profiles):
            try:
                dd = dataio.load_warmloop(fp, onto, pf)
                got = f"{pf} · 샘플 {len(dd.sample_ids)} · 축 {len(dd.y_terms)}"
                break
            except Exception:                                  # noqa: BLE001
                continue
        rows_e.append({"파일": f, "읽힌 내용": got,
                       "크기": f"{os.path.getsize(fp) // 1024} KB"})
    if rows_e:
        st.dataframe(rows_e, use_container_width=True, hide_index=True)
    else:
        st.caption("아직 없습니다.")


