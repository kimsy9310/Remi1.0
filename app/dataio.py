# -*- coding: utf-8 -*-
"""
웜루프 데이터 입출력 — 랩에서 채운 xlsx 를 모형이 먹는 (X, Y) 로 바꾼다.

입력 파일 규약(RM_warmloop_260720_fixed.xlsx 형식)
--------------------------------------------------
  recipes      sample_id · is_benchmark · ing_*  (배합, 배치 그램 또는 %)
  sensory      sample_id · rep · sens_*          (벤치마크 상대, -3..+3)
  dictionary   column_name -> ontology_id (ING.*)
  instrumental (선택) sample_id · instr_*

두 가지를 여기서 흡수한다.

1. **단위.** 배합은 배치 그램으로 적히는 일이 많다(쌀농축액 100, 해바라기유 15…).
   혼합물 모형은 합계가 고정된 데이터에서만 뜻이 있으므로 행마다 100 으로
   정규화한다. 원본 합계는 남겨서 되돌릴 수 있게 한다.

2. **컬럼 이름.** 관능 컬럼(sens_*)은 사람이 쓴 이름이고 모형은 L.* 를 쓴다.
   M 카드의 legacy_column 필드가 그 다리다 — 매핑을 코드에 박지 않고
   온톨로지에 둔다. 새 제품은 카드만 쓰면 되고 이 파일은 그대로다.
"""
from __future__ import annotations

import os
from dataclasses import dataclass, field

import numpy as np
import openpyxl


@dataclass
class WarmLoopData:
    names: list                      # 팔레트(필러 포함), ING.*
    X: np.ndarray                    # (n, p) 합계 100 으로 정규화
    Y: np.ndarray                    # (n, m) 관능, 벤치마크 상대
    y_terms: list                    # L.* 축 이름
    sample_ids: list
    benchmark_ids: list = field(default_factory=list)
    raw_totals: np.ndarray = None    # 정규화 전 행 합계
    unmapped_columns: list = field(default_factory=list)
    missing_axes: list = field(default_factory=list)
    notes: list = field(default_factory=list)

    def summary(self):
        L = [f"샘플 {len(self.sample_ids)}건 · 재료 {len(self.names)}종 · 관능축 {len(self.y_terms)}개"]
        if self.benchmark_ids:
            L.append(f"벤치마크: {', '.join(self.benchmark_ids)}")
        if self.unmapped_columns:
            L.append(f"온톨로지에 매핑 안 된 컬럼 {len(self.unmapped_columns)}개: "
                     f"{self.unmapped_columns[:5]}")
        if self.missing_axes:
            L.append(f"카드는 있으나 데이터에 없는 축: {self.missing_axes}")
        L += self.notes
        return "\n".join("  " + s for s in L)


def _rows(ws):
    it = ws.iter_rows(values_only=True)
    header = next(it)
    header = [None if h is None else str(h).strip() for h in header]
    out = []
    for r in it:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append(dict(zip(header, r)))
    return header, out


def _num(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def read_dictionary(wb):
    """column_name -> ING.* 매핑. 헤더 위치가 파일마다 조금씩 달라 훑어서 찾는다."""
    if "dictionary" not in wb.sheetnames:
        return {}
    m = {}
    for r in wb["dictionary"].iter_rows(values_only=True):
        cells = [None if c is None else str(c).strip() for c in r]
        if not cells or not cells[0]:
            continue
        col = cells[0]
        oid = next((c for c in cells[1:] if c and c.startswith("ING.")), None)
        if col.startswith("ing_"):
            m[col] = oid            # oid 가 None 일 수 있다(물 등) — 아래에서 처리
    return m


def load_warmloop(path, onto, profile, filler=None):
    """xlsx 한 개 → WarmLoopData. onto 는 V2Ontology."""
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    for need in ("recipes", "sensory"):
        if need not in wb.sheetnames:
            raise ValueError(f"'{need}' 시트가 없습니다. 시트: {wb.sheetnames}")

    filler = filler or onto.filler_of(profile)
    notes, unmapped = [], []

    # ---- 1) 관능 컬럼 -> L.* : M 카드의 legacy_column 이 다리다
    cards = onto._ref.load_cards(profile, onto.layers)
    col2term, term2col = {}, {}
    for c in cards:
        lc = c.get("legacy_column")
        if lc:
            col2term[lc] = c["term_id"]
            term2col[c["term_id"]] = lc
    if not col2term:
        raise ValueError(
            f"프로파일 {profile} 의 M 카드에 legacy_column 이 없습니다. "
            f"관능 컬럼을 L.* 로 잇는 다리가 없으면 데이터를 붙일 수 없습니다.")

    # ---- 2) 재료 컬럼 -> ING.*
    ing_map = read_dictionary(wb)
    _, rec_rows = _rows(wb["recipes"])
    ing_cols = [c for c in (rec_rows[0].keys() if rec_rows else []) if c and c.startswith("ing_")]
    resolved = {}
    for c in ing_cols:
        oid = ing_map.get(c)
        if not oid:
            # 사전에 없거나 비어 있는 경우: 물은 필러로, 나머지는 미매핑으로 남긴다
            if "water" in c.lower():
                oid = filler
                notes.append(f"'{c}' 를 필러 {filler} 로 해석했습니다(사전에 ID 없음).")
            else:
                unmapped.append(c)
                continue
        if oid not in onto.ingredients:
            unmapped.append(f"{c}->{oid}(온톨로지에 없음)")
            continue
        resolved.setdefault(oid, []).append(c)

    if not resolved:
        raise ValueError("배합 컬럼을 하나도 온톨로지 ID 로 잇지 못했습니다.")

    # ---- 3) 관능: rep 평균
    _, sen_rows = _rows(wb["sensory"])
    sen = {}
    for r in sen_rows:
        sid = r.get("sample_id")
        if not sid:
            continue
        sen.setdefault(str(sid).strip(), []).append(r)

    # 축을 고를 때는 "컬럼이 있는가" 가 아니라 "몇 건이나 채워졌는가" 를 본다.
    # 실제 데이터에서 creaminess 는 컬럼은 있지만 20건 중 1건만 채워져 있었다
    # (시트 자체 메모: "creaminess=신규(과거 데이터 없음)"). 이런 축을 Y 에 넣으면
    # 행 전체가 버려진다 — 축 하나 때문에 데이터셋이 사라진다.
    rec_ids = {str(r.get("sample_id")).strip() for r in rec_rows if r.get("sample_id")}
    usable = [s for s in sen if s in rec_ids]
    n_samples = len(usable)
    min_filled = max(4, int(round(0.5 * n_samples)))

    y_terms, missing, thin = [], [], []
    for c in cards:
        t = c["term_id"]
        lc = term2col.get(t)
        cnt = sum(1 for s in usable
                  for rr in sen[s] if _num(rr.get(lc)) is not None)
        if cnt == 0:
            missing.append(t)
        elif cnt < min_filled:
            thin.append((t, cnt))
        else:
            y_terms.append(t)
    for t, cnt in thin:
        notes.append(
            f"{t}: {cnt}/{n_samples} 건만 채워져 이번 학습에서 제외했습니다. "
            f"(축을 지운 것이 아니라 이 데이터로는 못 배웁니다 — 다음 배치에서 채우면 살아납니다)")
    if not y_terms:
        raise ValueError(
            "관능 시트에서 학습에 쓸 만큼 채워진 축을 찾지 못했습니다. "
            f"컬럼은 있으나 값이 부족한 축: {thin}")

    # ---- 4) 행 조립 (배합과 관능이 모두 있는 샘플만)
    names = sorted(resolved)
    if filler not in names:
        names.append(filler)
        notes.append(f"필러 {filler} 가 배합 컬럼에 없어 잔량으로 채웁니다.")

    X, Y, sids, benches, totals = [], [], [], [], []
    for r in rec_rows:
        sid = r.get("sample_id")
        if not sid:
            continue
        sid = str(sid).strip()
        if sid not in sen:
            continue
        row = np.zeros(len(names))
        for oid, cols in resolved.items():
            v = sum(_num(r.get(c)) or 0.0 for c in cols)
            row[names.index(oid)] = v
        tot = row.sum()
        if tot <= 0:
            notes.append(f"{sid}: 배합 합계가 0 이라 건너뜁니다.")
            continue

        yv, ok = np.zeros(len(y_terms)), True
        for k, t in enumerate(y_terms):
            vals = [_num(rr.get(term2col[t])) for rr in sen[sid]]
            vals = [v for v in vals if v is not None]
            if not vals:
                ok = False
                break
            yv[k] = float(np.mean(vals))
        if not ok:
            notes.append(f"{sid}: 관능 값이 비어 있어 건너뜁니다.")
            continue

        X.append(row / tot * 100.0)      # 혼합물 제약: 합계 100
        totals.append(tot)
        Y.append(yv)
        sids.append(sid)
        if str(r.get("is_benchmark", "")).strip().lower() in ("yes", "y", "true", "1"):
            benches.append(sid)

    if not X:
        raise ValueError("배합과 관능이 모두 채워진 샘플이 없습니다.")

    return WarmLoopData(
        names=names, X=np.array(X), Y=np.array(Y), y_terms=y_terms,
        sample_ids=sids, benchmark_ids=benches,
        raw_totals=np.array(totals), unmapped_columns=unmapped,
        missing_axes=missing, notes=notes)


def write_template(path, onto, profile, palette, sample_ids=None):
    """랩이 채울 빈 xlsx 를 만든다. 컬럼 이름이 곧 규약이므로 여기서 생성한다."""
    cards = onto._ref.load_cards(profile, onto.layers)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "recipes"
    ws.append(["sample_id", "is_benchmark", "notes"] + [f"ing_{g}" for g in palette])

    ws2 = wb.create_sheet("sensory")
    cols = []
    for c in cards:
        cols.append(c.get("legacy_column") or c["term_id"])
    ws2.append(["sample_id", "rep"] + cols)

    ws3 = wb.create_sheet("dictionary")
    ws3.append(["column_name", "real_ingredient", "ontology_id (ING.*)"])
    for g in palette:
        ws3.append([f"ing_{g}", "", g])

    ws4 = wb.create_sheet("README")
    for line in [
        "웜루프 데이터 템플릿",
        "",
        f"프로파일: {profile}",
        "",
        "1. recipes 시트에 배합을 적습니다. 단위는 배치 그램이든 % 든 상관없습니다 —",
        "   행마다 자동으로 100 으로 정규화합니다.",
        "2. sensory 시트는 벤치마크 대비 상대평가입니다. -3 ~ +3, 0 = 벤치마크와 같음.",
        "   벤치마크 자신은 전 축 0 으로 적고 recipes 에서 is_benchmark=yes 로 표시합니다.",
        "3. 같은 샘플을 여러 번 평가했으면 rep 을 1,2,3… 으로 늘려 적으세요. 평균을 씁니다.",
        "",
        "축 정의와 평가 요령:",
    ]:
        ws4.append([line])
    for c in cards:
        ws4.append([f"  {c.get('legacy_column') or c['term_id']}  ({c['term_id']})"])
        if c.get("evaluation_note"):
            ws4.append([f"      {c['evaluation_note']}"])

    if sample_ids:
        for sid in sample_ids:
            ws.append([sid])
            ws2.append([sid, 1])
    wb.save(path)
    return path
