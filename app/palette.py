# -*- coding: utf-8 -*-
"""
팔레트 표 로더 — data/palette.xlsx 를 읽어 모형이 쓰는 형태로 준다.

이 파일이 재료 범위의 **정본**이다. SQLite 가 아니라 xlsx 인 이유는 단순하다 —
이 프로젝트의 모든 큐레이션이 이미 엑셀로 되어 있고, 등급·메모·근거를 함께
보면서 고치는 일이라 표가 맞다. DB 는 실행 이력만 맡는다.

표 한 줄이 (프로파일 × 재료) 하나다:

    프로파일 · 슬롯 · 재료 · 온톨로지ID · 등급 · 하한 · 상한 · 범위근거 · 담당축 · 메모 · 확인

앱이 여기서 얻는 것
  bounds()   {ING.*: (lo, hi)}   제안의 상·하한이자 사전계수의 스케일
  palette()  [ING.*]             제외 등급을 뺀 재료 목록
  slots()    {슬롯: [ING.*]}      슬롯 하나가 변수 하나 — 중복 재료를 묶는 근거
"""
from __future__ import annotations

import os
from dataclasses import dataclass, field

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_PATH = os.path.join(os.path.dirname(_HERE), "data", "palette.xlsx")

COLS = ["프로파일", "슬롯", "재료", "온톨로지ID", "등급",
        "하한", "상한", "범위근거", "담당축", "메모", "확인"]

GRADES = ["필수", "권장", "옵션", "제한", "제외"]
USABLE = ("필수", "권장", "옵션")        # 제한은 경고와 함께 허용, 제외는 뺀다


def _f(v):
    if v is None or (isinstance(v, str) and not str(v).strip()):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


@dataclass
class PaletteTable:
    profile: str
    rows: list = field(default_factory=list)
    path: str = ""

    # ------------------------------------------------------------- 기본 조회
    def _usable(self, include_restricted=True):
        ok = set(USABLE) | ({"제한"} if include_restricted else set())
        return [r for r in self.rows if r["등급"] in ok]

    def palette(self, include_restricted=False, only_with_range=False):
        """모형에 넣을 재료 목록. 제외 등급은 빠진다."""
        rs = self._usable(include_restricted)
        if only_with_range:
            rs = [r for r in rs if r["하한"] is not None and r["상한"] is not None]
        seen, out = set(), []
        for r in rs:
            g = r["온톨로지ID"]
            if g not in seen:
                seen.add(g)
                out.append(g)
        return out

    def bounds(self, include_restricted=False):
        """{ING.*: (lo, hi)} — 둘 다 채워진 행만. adapter/propose 가 그대로 받는다."""
        out = {}
        for r in self._usable(include_restricted):
            lo, hi = r["하한"], r["상한"]
            if lo is None or hi is None or hi <= lo:
                continue
            out[r["온톨로지ID"]] = (lo, hi)
        return out

    def slots(self, include_restricted=False):
        out = {}
        for r in self._usable(include_restricted):
            out.setdefault(r["슬롯"] or "(미분류)", []).append(r["온톨로지ID"])
        return out

    def grade_of(self, ing_id):
        for r in self.rows:
            if r["온톨로지ID"] == ing_id:
                return r["등급"]
        return None

    def excluded(self):
        return [r["온톨로지ID"] for r in self.rows if r["등급"] == "제외"]

    def restricted(self):
        return [r["온톨로지ID"] for r in self.rows if r["등급"] == "제한"]

    # ------------------------------------------------------------- 상태 점검
    def missing_ranges(self):
        """쓸 재료인데 범위가 없는 것. 이 목록이 줄어들수록 제안을 믿을 수 있다."""
        out = []
        for r in self._usable():
            if r["하한"] is None or r["상한"] is None:
                out.append(r["온톨로지ID"])
        return out

    def bad_ranges(self):
        out = []
        for r in self._usable():
            lo, hi = r["하한"], r["상한"]
            if lo is not None and hi is not None and hi <= lo:
                out.append((r["온톨로지ID"], lo, hi))
        return out

    def no_axis(self):
        """담당축이 빈 재료 — 이 프로파일에서 아무 목표축도 못 움직인다."""
        return [r["온톨로지ID"] for r in self._usable() if not (r["담당축"] or "").strip()]

    def flagged(self):
        return [(r["온톨로지ID"], r["확인"]) for r in self.rows if (r["확인"] or "").strip()]

    def coverage(self, core_terms):
        """목표축마다 그 축을 움직이는 재료가 몇 개인지 (G2)."""
        out = {t: [] for t in core_terms}
        for r in self._usable():
            ax = (r["담당축"] or "")
            for t in core_terms:
                if t.split(".")[-1] in ax:
                    out[t].append(r["온톨로지ID"])
        return out

    def status(self):
        u = self._usable()
        return dict(rows=len(self.rows), usable=len(u),
                    with_range=len(u) - len(self.missing_ranges()),
                    missing=len(self.missing_ranges()),
                    excluded=len(self.excluded()),
                    flagged=len(self.flagged()))


# ------------------------------------------------------------------ 읽기/쓰기
def load(profile, path=None):
    p = path or DEFAULT_PATH
    if not os.path.exists(p):
        raise FileNotFoundError(
            f"팔레트 표가 없습니다: {p}\n"
            f"`python tools/build_palette.py` 로 만드세요.")
    wb = openpyxl.load_workbook(p, data_only=True)
    if "palette" not in wb.sheetnames:
        raise ValueError(f"'palette' 시트가 없습니다. 시트: {wb.sheetnames}")
    ws = wb["palette"]
    it = ws.iter_rows(values_only=True)
    hdr = [None if h is None else str(h).strip() for h in next(it)]
    miss = [c for c in ("프로파일", "온톨로지ID", "등급") if c not in hdr]
    if miss:
        raise ValueError(f"필수 열이 없습니다: {miss}. 있는 열: {hdr}")

    rows = []
    for r in it:
        d = dict(zip(hdr, r))
        if not d.get("온톨로지ID"):
            continue
        if str(d.get("프로파일") or "").strip() != profile:
            continue
        grade = str(d.get("등급") or "").strip()
        rows.append(dict(
            프로파일=profile,
            슬롯=str(d.get("슬롯") or "").strip(),
            재료=str(d.get("재료") or "").strip(),
            온톨로지ID=str(d["온톨로지ID"]).strip(),
            등급=grade if grade in GRADES else "옵션",
            하한=_f(d.get("하한")), 상한=_f(d.get("상한")),
            범위근거=str(d.get("범위근거") or "").strip(),
            담당축=str(d.get("담당축") or "").strip(),
            메모=str(d.get("메모") or "").strip(),
            확인=str(d.get("확인") or "").strip()))
    return PaletteTable(profile=profile, rows=rows, path=p)


def profiles_in(path=None):
    p = path or DEFAULT_PATH
    if not os.path.exists(p):
        return []
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb["palette"]
    it = ws.iter_rows(values_only=True)
    hdr = [None if h is None else str(h).strip() for h in next(it)]
    i = hdr.index("프로파일")
    seen = []
    for r in it:
        v = r[i]
        if v and str(v).strip() not in seen:
            seen.append(str(v).strip())
    return seen


def save_rows(profile, rows, path=None):
    """
    한 프로파일의 행을 표에 되쓴다. 다른 프로파일 행과 시트는 건드리지 않는다.
    앱의 편집 화면이 부르는 경로다.
    """
    p = path or DEFAULT_PATH
    wb = openpyxl.load_workbook(p)
    ws = wb["palette"]
    hdr = [None if c.value is None else str(c.value).strip() for c in ws[1]]

    keep = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c + 1).value for c in range(len(hdr))]
        d = dict(zip(hdr, vals))
        if str(d.get("프로파일") or "").strip() != profile:
            keep.append(vals)

    fills = {}
    for r in range(2, ws.max_row + 1):
        fills[r] = ws.cell(r, 1).fill.fgColor.rgb
    ws.delete_rows(2, ws.max_row)

    for vals in keep:
        ws.append(vals)
    for d in rows:
        ws.append([d.get(c, "") if d.get(c) is not None else "" for c in hdr])
    wb.save(p)
    return p
